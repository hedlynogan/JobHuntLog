import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Author: Ned Hogan 01/19/24

# Prompt the user for their name
user_name = input("Please enter your name for the EDD Job Hunt Log title: ")

# Define the columns for EDD job reporting requirements
columns = [
    'Week Ending Date', 'Employer Name', 'Type of Work Sought', 'Method of Contact',
    'Date of Contact', 'Position Applied', 'Contact Person', "Contact's Title",
    'Howto Contact ', 'Result', 'Follow-up Actions'
]

# Add a title to the spreadsheet using the user's name
current_date = datetime.now().strftime('%Y-%m-%d')
title = f"{user_name} Job Log - {current_date}"

# Create a DataFrame with the title and defined columns
# The title will be in the first row, and the column headers will be in the second row
# edd_job_reporting_log_with_title = pd.DataFrame(columns=[''])
title_df = pd.DataFrame({'A': [title]})
columns_df = pd.DataFrame(columns=columns)
edd_job_reporting_log_with_title = pd.concat([title_df, columns_df], ignore_index=True)

edd_file_name_with_title = f"{user_name}_Job_Log_{current_date}.xlsx"

# Using openpyxl for more control over the Excel file
wb = Workbook()
ws = wb.active
ws.title = "Job Log"
title_font = Font(bold=True, size=14)
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

ws.append([title])
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
ws['A1'].alignment = Alignment(horizontal='center')

# Append the column headers
ws.append(columns)

# Adjust column widths based on the length of column headers
for col, column_title in enumerate(columns, start=1):
    ws.column_dimensions[chr(64 + col)].width = len(column_title) + 2

# Save the workbook
wb.save(filename=edd_file_name_with_title)

# Output the file name
edd_file_name_with_title




